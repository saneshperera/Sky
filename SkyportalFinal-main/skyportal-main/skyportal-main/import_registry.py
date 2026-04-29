import os
import django
from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "skyportal.settings")
django.setup()

from teams.models import Team


FILE_NAME = "Agile Project Module UofW - Team Registry.xlsx"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


workbook = load_workbook(FILE_NAME, data_only=True)
sheet = workbook.active

created_count = 0
updated_count = 0

for row in sheet.iter_rows(min_row=2, values_only=True):
    department = clean(row[0])
    team_leader = clean(row[1])
    department_head = clean(row[2])
    team_name = clean(row[3])
    github_repo = clean(row[6])
    jira_board = clean(row[7])
    development_focus = clean(row[8])
    skills = clean(row[9])
    downstream_dependencies = clean(row[10])
    dependency_type = clean(row[11])
    software_owned = clean(row[12])
    slack_channels = clean(row[15])
    team_wiki = clean(row[18])

    if not team_name:
        continue

    team, created = Team.objects.update_or_create(
        team_name=team_name,
        defaults={
            "department": department,
            "team_leader": team_leader,
            "department_head": department_head,
            "team_type": dependency_type,
            "mission": development_focus,
            "responsibilities": software_owned,
            "skills": skills,
            "contact_email": "",
            "slack_channel": slack_channels,
            "code_repository": github_repo,
            "jira_board": jira_board,
            "wiki_page": team_wiki,
            "upstream_dependencies": "",
            "downstream_dependencies": downstream_dependencies,
        }
    )

    if created:
        created_count += 1
    else:
        updated_count += 1

print("Import complete.")
print(f"Created teams: {created_count}")
print(f"Updated teams: {updated_count}")
print(f"Total teams in database: {Team.objects.count()}")