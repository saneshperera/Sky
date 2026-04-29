from openpyxl import load_workbook

file_name = "Agile Project Module UofW - Team Registry.xlsx"

workbook = load_workbook(file_name)
sheet = workbook.active

print("Sheet name:", sheet.title)
print("Rows:", sheet.max_row)
print("Columns:", sheet.max_column)

print("\nColumn headings:")
for cell in sheet[1]:
    print(cell.column, cell.value)

print("\nFirst data row:")
for cell in sheet[2]:
    print(cell.column, cell.value)